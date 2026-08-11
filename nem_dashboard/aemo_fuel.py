"""Build daily NEM fuel totals directly from public AEMO source files."""

import csv
import io
import re
import shutil
from collections import defaultdict
from datetime import datetime, time, timedelta, timezone
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlsplit
from urllib.request import Request, urlopen
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


NEM_TIME = timezone(timedelta(hours=10), name='NEM')
SCADA_ARCHIVE = 'https://nemweb.com.au/Reports/Archive/Dispatch_SCADA/'
DEFAULT_GENERATION_REGISTER_URL = (
    'https://www.aemo.com.au/-/media/files/electricity/nem/planning_and_forecasting/'
    'generation_information/2026/nem-generation-information-july-2026.xlsx'
    '?rev=3455851f2bc945b7ab61c5ceed272992&sc_lang=en'
)
USER_AGENT = 'accidentalscientist.net NEM-dashboard (+https://accidentalscientist.net)'
OPERATING_STATUSES = {'In Service', 'In Commissioning', 'Announced Withdrawal'}


class FuelSourceError(RuntimeError):
    pass


def _download(url, destination, *, refresh=False, require_zip=False):
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists() and not refresh:
        if not require_zip:
            return destination
        try:
            with ZipFile(destination) as archive:
                if archive.testzip() is None:
                    return destination
        except BadZipFile:
            pass
        destination.unlink()

    temporary = destination.with_suffix(destination.suffix + '.part')
    try:
        request = Request(url, headers={'User-Agent': USER_AGENT})
        with urlopen(request, timeout=120) as response, temporary.open('wb') as output:
            shutil.copyfileobj(response, output, length=1024 * 1024)
        if require_zip:
            with ZipFile(temporary) as archive:
                corrupt = archive.testzip()
                if corrupt:
                    raise FuelSourceError(f'{destination.name} contains a corrupt member: {corrupt}')
        temporary.replace(destination)
    except (FuelSourceError, BadZipFile, HTTPError, URLError, TimeoutError, OSError) as exc:
        if temporary.exists():
            temporary.unlink()
        raise FuelSourceError(f'Could not download {url}: {exc}') from exc
    return destination


def register_cache_name(url):
    name = Path(urlsplit(url).path).name
    return name if name.lower().endswith('.xlsx') else 'nem-generation-information.xlsx'


def download_generation_register(url, cache_dir, *, refresh=False):
    return _download(url, Path(cache_dir) / register_cache_name(url), refresh=refresh)


def download_dispatch_scada(operating_date, cache_dir):
    filename = f'PUBLIC_DISPATCHSCADA_{operating_date:%Y%m%d}.zip'
    return _download(
        SCADA_ARCHIVE + filename,
        Path(cache_dir) / filename,
        require_zip=True,
    )


def classify_fuel(technology, detail='', gas_fuel=''):
    technology = (technology or '').strip().lower()
    detail = (detail or '').strip().lower()
    gas_fuel = (gas_fuel or '').strip().lower()

    if technology == 'coal':
        return 'Brown coal' if 'brown' in detail else 'Black coal'
    if technology == 'battery storage':
        return 'Battery'
    if technology == 'hydro':
        return 'Hydro'
    if technology == 'wind':
        return 'Wind'
    if technology == 'solar pv':
        return 'Solar'
    if 'biofuel' in gas_fuel or 'biomass' in gas_fuel or 'landfill' in gas_fuel:
        return 'Biomass'
    if technology == 'gas turbine':
        return 'Liquid Fuel' if 'diesel' in gas_fuel or 'fuel oil' in gas_fuel else 'Gas'
    if 'biomass' in detail:
        return 'Biomass'
    return 'Other'


def _split_duids(value):
    return [item for item in re.split(r'[,;\s]+', str(value or '').strip()) if item]


def load_generation_register(path):
    """Return DUID -> (state, display fuel), preferring a specific fuel over Other."""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook['Generator Information']
    except (OSError, KeyError, BadZipFile) as exc:
        raise FuelSourceError(f'Could not read generation register {path}: {exc}') from exc

    header_row = None
    headers = None
    for row_number, row in enumerate(sheet.iter_rows(max_row=20, values_only=True), start=1):
        values = [str(value).strip() if value is not None else '' for value in row]
        if 'DUID' in values and 'Technology Type' in values:
            header_row, headers = row_number, values
            break
    if not header_row:
        raise FuelSourceError('Generation register has no DUID/Technology Type header row.')

    columns = {name: headers.index(name) for name in (
        'DUID', 'Region', 'Technology Type', 'Technology Detail',
        'Gas Turbine Fuel Type', 'Commitment Status',
    )}
    registry = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        # Announced-withdrawal units still generate until their closure date,
        # and commissioning units can already appear in SCADA. The dispatch
        # record, not the planning label, determines whether they contributed.
        if str(row[columns['Commitment Status']] or '').strip() not in OPERATING_STATUSES:
            continue
        region = str(row[columns['Region']] or '').strip()
        state = region[:-1] if region.endswith('1') else region
        fuel = classify_fuel(
            row[columns['Technology Type']],
            row[columns['Technology Detail']],
            row[columns['Gas Turbine Fuel Type']],
        )
        for duid in _split_duids(row[columns['DUID']]):
            previous = registry.get(duid)
            if previous is None or (previous[1] == 'Other' and fuel != 'Other'):
                registry[duid] = (state, fuel)
    if not registry:
        raise FuelSourceError('Generation register contained no operating DUIDs.')
    return registry


def _mms_rows(binary_stream):
    headers = None
    reader = csv.reader(io.TextIOWrapper(binary_stream, encoding='utf-8-sig', errors='replace'))
    for row in reader:
        if len(row) < 4 or row[1:3] != ['DISPATCH', 'UNIT_SCADA']:
            continue
        if row[0] == 'I':
            headers = row[4:]
        elif row[0] == 'D' and headers:
            values = row[4:]
            if len(values) < len(headers):
                values.extend([''] * (len(headers) - len(values)))
            yield dict(zip(headers, values))


def aggregate_dispatch_scada(path, operating_date, registry):
    """Convert five-minute MW observations into daily MWh by state and fuel."""
    start = datetime.combine(operating_date, time.min, tzinfo=NEM_TIME)
    end = start + timedelta(days=1)
    observations = {}
    try:
        with ZipFile(path) as outer:
            for nested_info in outer.infolist():
                if not nested_info.filename.lower().endswith('.zip'):
                    continue
                with ZipFile(io.BytesIO(outer.read(nested_info))) as nested:
                    for csv_info in nested.infolist():
                        if not csv_info.filename.lower().endswith('.csv'):
                            continue
                        with nested.open(csv_info) as stream:
                            for row in _mms_rows(stream):
                                try:
                                    when = datetime.strptime(
                                        row['SETTLEMENTDATE'], '%Y/%m/%d %H:%M:%S'
                                    ).replace(tzinfo=NEM_TIME)
                                    value = float(row['SCADAVALUE'])
                                except (KeyError, TypeError, ValueError):
                                    continue
                                if start < when <= end:
                                    observations[(row.get('DUID', '').strip(), when)] = value
    except BadZipFile as exc:
        raise FuelSourceError(f'Invalid dispatch SCADA archive {path}: {exc}') from exc

    totals = defaultdict(float)
    unknown = defaultdict(float)
    mapped_mwh = 0.0
    unknown_mwh = 0.0
    for (duid, _when), value in observations.items():
        # Generation fuel mix is gross output. Charging and pumping load are
        # system demand and must not subtract from the day's generated energy.
        energy = max(value, 0.0) / 12.0
        if not energy:
            continue
        identity = registry.get(duid)
        if identity:
            totals[identity] += energy
            mapped_mwh += energy
        else:
            unknown[duid] += energy
            unknown_mwh += energy

    denominator = mapped_mwh + unknown_mwh
    coverage = mapped_mwh / denominator if denominator else 0.0
    return {
        'totals': dict(totals),
        'coverage': coverage,
        'mapped_mwh': mapped_mwh,
        'unknown_mwh': unknown_mwh,
        'unknown': sorted(unknown.items(), key=lambda item: item[1], reverse=True),
        'interval_records': len(observations),
    }
